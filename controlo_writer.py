"""
controlo_writer.py — Registo unificado de deals confirmados
===========================================================
Regista no ficheiro de controlo (estrutura igual ao BoxMovers2026_i, folha DEALS
/ Table13, + coluna NEGÓCIO) todos os deals a partir de "Encomenda Confirmada".
1 linha por SKU. Fontes: simulador (produto/custo — re-lookup ao índice se faltar),
o próprio deal (cliente/qty/preço/datas) e o campo Negócio do deal.

Local — só escreve quando a app corre nesta máquina (na cloud faz no-op seguro).
"""
import re
import json
import datetime
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

# Reutiliza os helpers e mapas de colunas do writer BoxMovers (mesma Table13)
from boxmovers_writer import (
    _deal_date, _ean, _num_part, INP, PROD, TEMPLATE_ROW, STATIC_PRODUCT_COLS,
)

try:
    from config import CONTROLO_FILE, NEGOCIOS
except Exception:
    CONTROLO_FILE = Path(__file__).parent.parent / "Controlo_Deals_2026.xlsx"
    NEGOCIOS = ["Box Movers", "African Markets Wholesale", "Franchising"]

NCOL_F      = 63    # colunas A-BK (com fórmulas/template)
NEGOCIO_COL = 64    # coluna BL — NEGÓCIO
UN_COL      = 16    # P — XLOOKUP a Query1 (Deep Info removido) → limpar
REGISTRY    = Path(str(CONTROLO_FILE) + ".registered.json")


def _load_reg() -> set:
    try:
        return set(json.loads(REGISTRY.read_text(encoding="utf-8")))
    except Exception:
        return set()


def is_registered(deal_id) -> bool:
    return str(deal_id) in _load_reg()


def _mark(deal_id):
    reg = _load_reg()
    reg.add(str(deal_id))
    try:
        REGISTRY.parent.mkdir(parents=True, exist_ok=True)
        REGISTRY.write_text(json.dumps(sorted(reg)), encoding="utf-8")
    except Exception:
        pass


def _enrich(data: dict, sku) -> dict:
    """Garante que produto/custo vêm do simulador — re-consulta o índice se faltar."""
    ok = data.get("ean") and str(data.get("ean")).lower() != "nan" \
        and data.get("name") and (data.get("pcl") or data.get("ufc_raw"))
    if ok:
        return data
    try:
        from sku_lookup import build_cache
        e = build_cache().get(str(sku)) or {}
        if not e:
            return data
        merged = dict(e)
        # dados do deal têm prioridade quando existem e são válidos
        for k, v in (data or {}).items():
            if v not in (None, "", "nan"):
                merged[k] = v
        return merged
    except Exception:
        return data


def append_deal(deal: dict, business: str = "", status_label: str = "CONCLUÍDO",
                force: bool = False) -> tuple:
    """Acrescenta 1 linha por SKU ao ficheiro de controlo. Devolve (n_linhas, msg)."""
    sd = deal.get("skus_detail") or deal.get("_skus_detail") or {}
    if not sd:
        return 0, "Deal sem SKUs — nada a registar."
    did = deal.get("deal_id")
    if not force and did and is_registered(did):
        return 0, f"{did} já tinha sido registado no controlo (ignorado)."

    fp = Path(CONTROLO_FILE)
    if not fp.exists():
        return 0, f"Ficheiro de controlo não encontrado: {fp.name}"

    wb = openpyxl.load_workbook(str(fp))
    ws = wb["DEALS"]
    tbl = ws.tables["Table13"]
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", tbl.ref)
    start = int(m.group(4)) + 1

    # separador da DEAL DATE (ex.: "Jan´26")
    sep = "´"
    for rr in range(3, int(m.group(4)) + 1):
        v = ws.cell(row=rr, column=1).value
        if isinstance(v, str):
            mm = re.match(r"[A-Za-z]{3}(.)\d{2}", v)
            if mm:
                sep = mm.group(1); break

    templ = {c: ws.cell(row=TEMPLATE_ROW, column=c) for c in range(1, NCOL_F + 1)}
    is_f = {c: isinstance(templ[c].value, str) and templ[c].value.startswith("=")
            for c in range(1, NCOL_F + 1)}

    ddate = _deal_date(deal.get("created_at") or datetime.date.today().isoformat(), sep)
    notas = f'{did or ""} | {deal.get("status", "")}'

    r = start
    n = 0
    for sku, item in sd.items():
        data = _enrich(item.get("data") or {}, sku)
        for c in range(1, NCOL_F + 1):
            src = templ[c]
            dst = ws.cell(row=r, column=c)
            dst._style = copy(src._style)
            if is_f[c]:
                dst.value = Translator(src.value, origin=src.coordinate)\
                    .translate_formula(f"{get_column_letter(c)}{r}")
        # inputs (deal)
        ws.cell(r, INP["DEAL_DATE"]).value = ddate
        ws.cell(r, INP["STATUS"]).value    = status_label
        ws.cell(r, INP["NOTAS"]).value     = notas
        ws.cell(r, INP["CLIENT"]).value    = deal.get("client")
        ws.cell(r, INP["SKU"]).value       = int(sku) if str(sku).isdigit() else sku
        ws.cell(r, INP["QTY"]).value       = int(item.get("qty") or 1)
        ws.cell(r, INP["PV_WRT"]).value    = float(item.get("pvp") or 0)
        ws.cell(r, INP["APOIO"]).value     = float(item.get("so_neg") or 0)
        ws.cell(r, INP["EST_PAG"]).value   = "ABERTO"
        ws.cell(r, INP["NET_COST"]).value  = float(item.get("fc_final") or data.get("ufc_raw") or data.get("pcl") or 0)
        ws.cell(r, INP["TAXAS"]).value     = float(data.get("eis_total") or 0)
        ws.cell(r, INP["REBATES"]).value   = float(data.get("cgf_reb") or 0)
        ws.cell(r, INP["COMMENTS"]).value  = float(data.get("cgf_com") or 0)
        # produto (simulador)
        if data.get("brand"): ws.cell(r, PROD["BRAND"]).value = data["brand"]
        if data.get("ean"):   ws.cell(r, PROD["EAN"]).value   = _ean(data["ean"])
        if data.get("name"):  ws.cell(r, PROD["DESC"]).value  = data["name"]
        if data.get("cat"):
            ws.cell(r, PROD["DESC_CAT"]).value = data["cat"]
            ws.cell(r, PROD["CAT"]).value      = _num_part(data["cat"])
        # negócio
        ws.cell(r, NEGOCIO_COL).value = business or ""
        r += 1
        n += 1

    tbl.ref = f"A2:{get_column_letter(NEGOCIO_COL)}{start + n - 1}"
    if tbl.autoFilter is not None:
        tbl.autoFilter.ref = tbl.ref     # manter alinhado (senão o Excel acusa erro)
    for i, tc in enumerate(tbl.tableColumns, start=1):
        if i in (list(STATIC_PRODUCT_COLS) + [UN_COL]) and getattr(tc, "calculatedColumnFormula", None) is not None:
            tc.calculatedColumnFormula = None

    wb.save(str(fp))
    if did:
        _mark(did)
    return n, f"{n} linha(s) registadas no controlo (linhas {start}–{start + n - 1})."
