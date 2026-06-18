"""
boxmovers_writer.py — Regista uma proposta adjudicada no BoxMovers (TESTE)
==========================================================================
Escreve 1 linha por SKU na folha DEALS do BoxMovers2026_TESTE.xlsx:
- inputs (DEAL DATE, STATUS, CLIENT, SKU, QTY, PV WRT, NET COST, TAXAS, ...)
- campos de produto (BRAND/EAN/DESCRIÇÃO/CAT) a partir do skus_detail do deal
- restantes colunas mantêm as fórmulas da tabela (calculam no Excel)
Estende a Table13 e remove a fórmula de coluna calculada nas colunas de produto
preenchidas (para os valores não serem revertidos pelo Excel).
"""
import re
import json
import datetime
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

# Caminho do ficheiro BoxMovers:
#  - TEST_MODE → ficheiro de teste (fora do OneDrive: gravações + rápidas)
#  - produção  → secret/env BOXMOVERS_PATH (se definido); senão cai no de teste.
# Na cloud (Railway) o ficheiro não existe → append_deal faz no-op seguro.
_TEST_BM = Path(r"C:\Users\SOUSAVitorAraujo\BM_Teste\BoxMovers2026_TESTE.xlsx")
try:
    from config import TEST_MODE as _TEST_MODE, _get_secret as _cfg_secret
    _PROD_BM = _cfg_secret("BOXMOVERS_PATH", "")
    TEST_BM = _TEST_BM if _TEST_MODE else (Path(_PROD_BM) if _PROD_BM else _TEST_BM)
except Exception:
    TEST_BM = _TEST_BM
TEMPLATE_ROW = 3
NCOL = 63
PT = {1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
      7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"}

# colunas de input (1-based)
INP = dict(DEAL_DATE=1, STATUS=2, NOTAS=12, CLIENT=13, SKU=26, QTY=29, PV_WRT=30,
           APOIO=31, EST_PAG=33, NET_COST=34, TAXAS=36, REBATES=51, COMMENTS=53)
# colunas de produto preenchidas a partir do skus_detail.data
PROD = dict(CAT=17, DESC_CAT=18, BRAND=25, EAN=27, DESC=28)
STATIC_PRODUCT_COLS = [17, 18, 25, 27, 28]

# Registo anti-duplicados (deals já escritos no BoxMovers)
REGISTRY = TEST_BM.parent / "bm_registered.json"


def _load_registry() -> set:
    try:
        return set(json.loads(REGISTRY.read_text(encoding="utf-8")))
    except Exception:
        return set()


def is_registered(deal_id) -> bool:
    return str(deal_id) in _load_registry()


def _mark_registered(deal_id):
    reg = _load_registry()
    reg.add(str(deal_id))
    try:
        REGISTRY.parent.mkdir(parents=True, exist_ok=True)
        REGISTRY.write_text(json.dumps(sorted(reg)), encoding="utf-8")
    except Exception:
        pass


def _num_part(v):
    if v is None:
        return None
    m = re.match(r"\s*(\d+)", str(v))
    return int(m.group(1)) if m else None


def _ean(v):
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return int(s) if s.isdigit() and not s.startswith("0") else (s or None)


def _deal_date(created, sep="´"):
    try:
        d = datetime.date.fromisoformat(str(created)[:10])
        return f"{PT[d.month]}{sep}{str(d.year)[2:]}"
    except Exception:
        return None


def append_deal(deal: dict, file_path=TEST_BM, status_label="PO", force=False) -> tuple[int, str]:
    """Acrescenta as linhas do deal ao BoxMovers de teste. Devolve (n_linhas, msg).
    Anti-duplicados: se o deal já foi registado, não escreve (a menos de force=True)."""
    sd = deal.get("skus_detail") or deal.get("_skus_detail") or {}
    if not sd:
        return 0, "Proposta sem SKUs — nada a registar."
    _did = deal.get("deal_id")
    if not force and _did and is_registered(_did):
        return 0, f"{_did} já tinha sido registado no BoxMovers (ignorado)."

    file_path = Path(file_path)
    if not file_path.exists():
        return 0, f"Ficheiro não encontrado: {file_path.name}"

    wb = openpyxl.load_workbook(str(file_path))
    ws = wb["DEALS"]
    tbl = ws.tables["Table13"]
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", tbl.ref)
    start = int(m.group(4)) + 1

    # separador da DEAL DATE a partir de uma linha existente
    sep = "´"
    for rr in range(3, int(m.group(4)) + 1):
        v = ws.cell(row=rr, column=1).value
        if isinstance(v, str):
            mm = re.match(r"[A-Za-z]{3}(.)\d{2}", v)
            if mm:
                sep = mm.group(1)
                break

    templ = {c: ws.cell(row=TEMPLATE_ROW, column=c) for c in range(1, NCOL + 1)}
    is_f = {c: isinstance(templ[c].value, str) and templ[c].value.startswith("=") for c in range(1, NCOL + 1)}

    ddate = _deal_date(deal.get("created_at") or datetime.date.today().isoformat(), sep)
    notas = f'{deal.get("deal_id", "")} | {deal.get("status", "")}'

    r = start
    n = 0
    for sku, item in sd.items():
        data = item.get("data") or {}
        # estilo + fórmulas
        for c in range(1, NCOL + 1):
            src = templ[c]
            dst = ws.cell(row=r, column=c)
            dst._style = copy(src._style)
            if is_f[c]:
                dst.value = Translator(src.value, origin=src.coordinate)\
                    .translate_formula(f"{get_column_letter(c)}{r}")
        # inputs
        ws.cell(r, INP["DEAL_DATE"]).value = ddate
        ws.cell(r, INP["STATUS"]).value    = status_label
        ws.cell(r, INP["NOTAS"]).value     = notas
        ws.cell(r, INP["CLIENT"]).value    = deal.get("client")
        ws.cell(r, INP["SKU"]).value       = int(sku) if str(sku).isdigit() else sku
        ws.cell(r, INP["QTY"]).value       = int(item.get("qty") or 1)
        ws.cell(r, INP["PV_WRT"]).value    = float(item.get("pvp") or 0)
        ws.cell(r, INP["APOIO"]).value     = 0.0
        ws.cell(r, INP["EST_PAG"]).value   = "ABERTO"
        ws.cell(r, INP["NET_COST"]).value  = float(item.get("fc_final") or data.get("ufc_raw") or 0)
        ws.cell(r, INP["TAXAS"]).value     = float(data.get("eis_total") or 0)
        ws.cell(r, INP["REBATES"]).value   = float(data.get("cgf_reb") or 0)
        ws.cell(r, INP["COMMENTS"]).value  = float(data.get("cgf_com") or 0)
        # produto (do skus_detail.data)
        if data.get("brand"): ws.cell(r, PROD["BRAND"]).value = data["brand"]
        if data.get("ean"):   ws.cell(r, PROD["EAN"]).value   = _ean(data["ean"])
        if data.get("name"):  ws.cell(r, PROD["DESC"]).value  = data["name"]
        if data.get("cat"):
            ws.cell(r, PROD["DESC_CAT"]).value = data["cat"]
            ws.cell(r, PROD["CAT"]).value      = _num_part(data["cat"])
        r += 1
        n += 1

    tbl.ref = f"A2:BK{start + n - 1}"
    # remover fórmula de coluna calculada nas colunas de produto preenchidas
    for i, tc in enumerate(tbl.tableColumns, start=1):
        if i in STATIC_PRODUCT_COLS and getattr(tc, "calculatedColumnFormula", None) is not None:
            tc.calculatedColumnFormula = None

    wb.save(str(file_path))
    if _did:
        _mark_registered(_did)
    return n, f"{n} linha(s) registadas no BoxMovers (linhas {start}–{start + n - 1})."
