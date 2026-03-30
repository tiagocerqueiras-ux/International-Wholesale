"""
Transport Lookup — Simulador Logístico
======================================
Compara tarifas de 6 transportadoras para envios B2B.
Fonte: Simulador_Exportacao_V2.26 - B2B.xlsx
"""

import json
from pathlib import Path

from config import TRANSPORT_FILE, TRANSPORT_CACHE, CACHE_DIR

CARRIERS = ["DSV", "SCHENKER", "T&W", "RHENUS", "RANGEL", "PORTIR"]

# Fuel surcharges — fraction of freight (most recent data)
FUEL_RATES = {
    "DSV":      0.07,
    "SCHENKER": 0.0,
    "T&W":      0.0,
    "RHENUS":   0.0253,
    "RANGEL":   0.055,
    "PORTIR":   0.0,
}

# Insurance data: pct = fraction of base, min = minimum EUR, base = "cargo" or "cargo+freight"
INSURANCE = {
    "DSV":      {"pct": 0.0025, "min": 25.0, "base": "cargo+freight"},
    "SCHENKER": None,   # not offered
    "T&W":      "included",
    "RHENUS":   {"pct": 0.0022, "min": 30.0, "base": "cargo"},
    "RANGEL":   {"pct": 0.0030, "min": 30.0, "base": "cargo"},
    "PORTIR":   {"pct": 0.0025, "min": 25.0, "base": "cargo"},
}

# Map carrier sheet names to canonical names
_SHEET_CARRIER = {
    "DSV":      "DSV",
    "SCHENKER": "SCHENKER",
    "T&W":      "T&W",
    "RHENUS":   "RHENUS",
    "RANGEL":   "RANGEL",
    "PORTIR":   "PORTIR",
}

# TT sheet column map: col_index -> (carrier, mode)
_TT_COLS = {
    1: ("SCHENKER", "LTL"),
    2: ("DSV",      "LTL"),
    3: ("DSV",      "FTL"),
    4: ("T&W",      "LTL"),
    5: ("T&W",      "FTL"),
    6: ("RHENUS",   "LTL"),
    7: ("RANGEL",   "LTL"),
}

_DEP_COLS = {
    1: ("SCHENKER", "small"),
    2: ("SCHENKER", "large"),
    3: ("DSV",      "LTL"),
    4: ("DSV",      "FTL"),
    5: ("T&W",      "STD"),
    6: ("RHENUS",   "LTL"),
    7: ("RANGEL",   "LTL"),
}


def build_transport_cache() -> dict:
    """Le o Excel e constroi o cache em memoria + guarda em JSON."""
    import openpyxl
    wb = openpyxl.load_workbook(str(TRANSPORT_FILE), read_only=True, data_only=True)

    destinations: dict = {}
    country_cps: dict  = {}

    # Tarifas por carrier
    for sheet_name, carrier in _SHEET_CARRIER.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header
            country = row[0]
            cp      = row[1]
            c_cp    = str(row[2]) if row[2] else ""
            if not country or not c_cp:
                continue
            prices = []
            for col_idx in range(3, 36):  # columns 4-36 = pallets 1-33
                val = row[col_idx] if col_idx < len(row) else None
                if val is None or val == "No Service":
                    prices.append(None)
                else:
                    try:
                        prices.append(round(float(val), 4))
                    except (TypeError, ValueError):
                        prices.append(None)
            # Pad to 33 if shorter
            while len(prices) < 33:
                prices.append(None)
            prices = prices[:33]

            if c_cp not in destinations:
                destinations[c_cp] = {
                    "country": str(country),
                    "cp":      cp,
                }
                ctry_key = str(country)
                if ctry_key not in country_cps:
                    country_cps[ctry_key] = []
                if cp not in country_cps[ctry_key]:
                    country_cps[ctry_key].append(cp)

            destinations[c_cp][carrier] = prices

    # Transit times
    tt_data: dict = {}
    if "TT" in wb.sheetnames:
        ws_tt = wb["TT"]
        for i, row in enumerate(ws_tt.iter_rows(values_only=True)):
            if i == 0:
                continue
            dest = str(row[0]) if row[0] else ""
            if not dest:
                continue
            tt_data[dest] = {}
            for col_idx, (carrier, mode) in _TT_COLS.items():
                val = row[col_idx] if col_idx < len(row) else None
                key = f"{carrier}_{mode}"
                if val and val != "No Service":
                    try:
                        tt_data[dest][key] = int(val)
                    except (TypeError, ValueError):
                        pass

    # Departure frequencies
    dep_data: dict = {}
    if "DEPARTURES" in wb.sheetnames:
        ws_dep = wb["DEPARTURES"]
        for i, row in enumerate(ws_dep.iter_rows(values_only=True)):
            if i == 0:
                continue
            dest = str(row[0]) if row[0] else ""
            if not dest:
                continue
            dep_data[dest] = {}
            for col_idx, (carrier, mode) in _DEP_COLS.items():
                val = row[col_idx] if col_idx < len(row) else None
                key = f"{carrier}_{mode}"
                if val and val != "No Service":
                    dep_data[dest][key] = str(val)

    wb.close()

    # Sort country CPs
    for ctry in country_cps:
        try:
            country_cps[ctry] = sorted(country_cps[ctry], key=lambda x: (str(x).zfill(10)))
        except Exception:
            pass

    cache = {
        "version":      2,
        "carriers":     CARRIERS,
        "destinations": destinations,
        "tt":           tt_data,
        "departures":   dep_data,
        "countries":    sorted(country_cps.keys()),
        "country_cps":  country_cps,
    }

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    with open(TRANSPORT_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, separators=(",", ":"))

    return cache


def load_transport_cache() -> dict:
    """Carrega o cache. Reconstroi automaticamente se nao existir."""
    if not TRANSPORT_CACHE.exists():
        return build_transport_cache()
    try:
        with open(TRANSPORT_CACHE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return build_transport_cache()


def get_countries(cache: dict) -> list:
    return cache.get("countries", [])


def get_cps_for_country(country: str, cache: dict) -> list:
    return cache.get("country_cps", {}).get(country, [])


def get_quote(
    c_cp: str,
    n_pallets: int,
    cargo_value: float = 0.0,
    include_insurance: bool = False,
    cache: dict = None,
) -> list:
    """
    Calcula cotacao para todos os carriers disponiveis para um destino.
    c_cp: chave do destino (ex: 'Espanha28')
    n_pallets: 1-33
    cargo_value: valor da carga em EUR (para seguro)
    include_insurance: True = adicionar seguro ao total
    Devolve lista de dicts ordenada por total_eur asc.
    """
    if cache is None:
        cache = load_transport_cache()

    dest = cache.get("destinations", {}).get(c_cp)
    if not dest:
        return []

    idx     = max(0, min(n_pallets - 1, 32))  # 0-based index for pallets
    tt_map  = cache.get("tt", {}).get(c_cp, {})
    dep_map = cache.get("departures", {}).get(c_cp, {})

    results = []
    for carrier in CARRIERS:
        prices = dest.get(carrier)
        if not prices:
            continue
        base_price = prices[idx]
        if base_price is None:
            continue

        freight = round(base_price, 2)
        fuel    = round(freight * FUEL_RATES.get(carrier, 0.0), 2)
        total   = round(freight + fuel, 2)

        # Insurance
        ins_label = ""
        ins_value = 0.0
        ins_info  = INSURANCE.get(carrier)
        if ins_info == "included":
            ins_label = "Incluido no frete"
        elif ins_info is None:
            ins_label = "Nao disponivel"
        elif isinstance(ins_info, dict) and include_insurance and cargo_value > 0:
            base = cargo_value
            if ins_info["base"] == "cargo+freight":
                base = cargo_value + freight * 1.10
            ins_value = max(ins_info["min"], round(base * ins_info["pct"], 2))
            ins_label = f"{ins_value:.2f} EUR"
            total = round(total + ins_value, 2)

        # Best transit time (prefer LTL)
        tt_days = None
        for mode in ["LTL", "STD", "FTL"]:
            key = f"{carrier}_{mode}"
            if key in tt_map:
                tt_days = tt_map[key]
                break

        # Departures
        departure = ""
        for mode in ["LTL", "small", "STD"]:
            key = f"{carrier}_{mode}"
            if key in dep_map:
                departure = dep_map[key]
                break

        results.append({
            "carrier":   carrier,
            "freight":   freight,
            "fuel":      fuel,
            "insurance": ins_value,
            "ins_label": ins_label,
            "total":     total,
            "tt_days":   tt_days,
            "departure": departure,
        })

    results.sort(key=lambda x: x["total"])
    return results
