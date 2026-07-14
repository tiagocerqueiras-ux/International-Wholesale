"""
import_purchase_history.py — histórico de compras African Markets
==================================================================
Lê os ficheiros de encomendas African Markets e gera
data/purchase_history.json (commitado, lido na cloud), para dar
referência de preço por cliente+SKU na construção da proposta.

Uso:  py import_purchase_history.py [--push]
  --push : commit+push do snapshot só se mudar (Railway redeploya).
"""
import sys
import re
import json
import shutil
import tempfile
import warnings
import unicodedata
import subprocess
import datetime as _dt
from pathlib import Path

warnings.simplefilter("ignore")

_BASE = (Path.home() / "Worten"
         / "B2B Business Unit - PART.& INT. EXP. AND TRAD. - Documents"
         / "ÁREAS DE NEGÓCIO" / "AFRICAN MARKET")

# (ficheiro, folhas de encomendas a ler)
FILES = [
    (_BASE / "Encomendas_FRANCHISING CV 2026_Cris.xlsx", ["VASCONCELOS LOPES", "TECNOLAR"]),
    (_BASE / "Encomendas_AfricanMarkets 2026_Cris.xlsx", ["Encomendas"]),
]
OUT = Path(__file__).parent / "data" / "purchase_history.json"


def norm(s: str) -> str:
    s = str(s or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _open(p: Path):
    import openpyxl
    try:
        return openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except PermissionError:
        t = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); t.close()
        shutil.copy2(str(p), t.name)
        return openpyxl.load_workbook(t.name, read_only=True, data_only=True)


def _num(v):
    try:
        return float(str(v).replace("€", "").replace(",", ".").strip())
    except (TypeError, ValueError):
        return None


def _date_str(v):
    if isinstance(v, _dt.datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    return s[:10] if s else ""


def _find_header(ws):
    """Devolve (linha_1based, headers_lower) da linha que tem 'sku' e 'cliente'."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        low = [str(c).strip().lower() if c is not None else "" for c in row]
        if "sku" in low and "cliente" in low:
            return i, low
    return None, None


def _col_map(headers):
    def find(*keys):
        for i, h in enumerate(headers):
            if any(k in h for k in keys):
                return i
        return None
    return {
        "date":   0,   # em ambos os ficheiros a data é sempre a 1.ª coluna
        "status": find("estado"),
        "order":  find("código encomenda", "codigo encomenda"),
        "client": find("cliente"),
        "sku":    find("sku"),
        "desc":   find("descri"),
        "qty":    find("qty", "quantidade"),
        "price":  find("preço", "preco"),   # 'valor total' não contém 'preço'
    }


def build():
    records = []
    for path, sheets in FILES:
        if not path.exists():
            print(f"  AVISO: não existe {path.name}")
            continue
        wb = _open(path)
        for sh in sheets:
            if sh not in wb.sheetnames:
                print(f"  AVISO: folha '{sh}' não existe em {path.name}")
                continue
            ws = wb[sh]
            hrow, headers = _find_header(ws)
            if not hrow:
                print(f"  AVISO: cabeçalho não encontrado em {path.name}/{sh}")
                continue
            cm = _col_map(headers)
            for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                def g(key):
                    idx = cm.get(key)
                    return row[idx] if (idx is not None and idx < len(row)) else None
                sku = g("sku")
                price = _num(g("price"))
                client = g("client")
                if sku is None or price is None or not client:
                    continue
                sku_s = str(sku).strip()
                if sku_s.endswith(".0"):
                    sku_s = sku_s[:-2]
                if not sku_s or sku_s.lower() == "nan":
                    continue
                records.append({
                    "client":      str(client).strip(),
                    "client_norm": norm(client),
                    "sku":         sku_s,
                    "price":       round(price, 4),
                    "qty":         _num(g("qty")),
                    "date":        _date_str(g("date")),
                    "status":      norm(g("status")),
                    "order":       str(g("order") or "").strip(),
                    "desc":        str(g("desc") or "").strip()[:60],
                    "src":         f"{path.name}/{sh}",
                })
        wb.close()

    clients = sorted(set(r["client_norm"] for r in records))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"records": records, "clients": clients},
                              ensure_ascii=False), encoding="utf-8")
    return records, clients


def _git(*a):
    return subprocess.run(["git", *a], cwd=str(Path(__file__).parent),
                          capture_output=True, text=True)


def _commit_push(n):
    rel = "data/purchase_history.json"
    if _git("diff", "--quiet", "--", rel).returncode == 0:
        print("Snapshot igual — sem commit/push."); return
    _git("add", rel)
    c = _git("commit", "-m", f"Auto: atualiza historico de compras African Markets ({n} linhas)")
    if c.returncode != 0:
        print("git commit:", (c.stdout + c.stderr).strip()); return
    p = _git("push", "origin", "main")
    print("Publicado." if p.returncode == 0 else f"push falhou: {p.stderr.strip()}")


if __name__ == "__main__":
    recs, clients = build()
    import collections
    by_client = collections.Counter(r["client_norm"] for r in recs)
    print(f"Escrito: {OUT}  ({len(recs)} linhas · {len(clients)} clientes)")
    for c, n in by_client.most_common():
        print(f"  {c[:40]:40} {n} linhas")
    if "--push" in sys.argv:
        _commit_push(len(recs))
