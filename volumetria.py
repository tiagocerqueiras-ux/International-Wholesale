"""
volumetria.py — Unidades por palete por SKU (Fase 1, Passo 3)
=============================================================
Lê o ficheiro "Informação Volumetria e HS Codes.xlsx" (folha VOLUMETRIA):
  cod_product (SKU) -> volume unitário (cm³).
unidades_por_palete = floor(capacidade_palete / volume_unitário)
paletes = teto(quantidade / unidades_por_palete)

A capacidade da palete (default 1.8 m³) é configurável. Lookup em cache local.
"""
import os
import json
import math
from pathlib import Path

# Localização do ficheiro (biblioteca sincronizada INC - BI & Analytics) + fallbacks
_CANDIDATES = [
    Path.home() / "Worten" / "INC - BI & Analytics - Documents" / "1. BI & Analytics"
        / "3. Desenvolvimentos" / "Franchising - Worten Cabo Verde" / "Informação Volumetria e HS Codes.xlsx",
    Path.home() / "Downloads" / "Informação Volumetria e HS Codes.xlsx",
]
CACHE = Path(__file__).resolve().parent / ".cache" / "volumetria.json"

# Capacidade de uma EURO-PALETE STANDARD: 1,2 m × 0,8 m × 1,8 m (altura) = 1,728 m³
PALLET_CAP_CM3 = 1_728_000

# Colunas da folha VOLUMETRIA (0-based)
C_SKU = 5     # VOLUMETRIA[cod_product]
C_VOL = 9     # [Sumval_unit_size_volume]  (cm³ = A×L×P)
C_H, C_L, C_W = 10, 11, 12  # altura, comprimento, largura (cm) — fallback do volume


def pallets_for(qty, units_per_pallet):
    """Nº de paletes = teto(qtd / unidades_por_palete). None se inválido."""
    try:
        upp = float(units_per_pallet)
        q = float(qty)
        if upp <= 0 or q <= 0:
            return None
        return math.ceil(q / upp)
    except (TypeError, ValueError):
        return None


def units_per_pallet_from_volume(unit_vol_cm3, cap_cm3=PALLET_CAP_CM3):
    try:
        v = float(unit_vol_cm3)
        if v < 5:          # volume implausível/em falta → deixar para manual
            return None
        return max(1, min(9999, math.floor(cap_cm3 / v)))
    except (TypeError, ValueError):
        return None


def _find_file():
    for c in _CANDIDATES:
        if c.exists():
            return c
    # fallback: procurar pelo nome
    for root in [Path.home() / "Worten", Path.home() / "Downloads"]:
        if not root.exists():
            continue
        for dp, dns, fns in os.walk(root, onerror=lambda e: None):
            dns[:] = [d for d in dns if not d.startswith(".")]
            for f in fns:
                if "volumetria e hs" in f.lower() and f.lower().endswith(".xlsx"):
                    return Path(dp) / f
    return None


def load_lookup(force: bool = False, cap_cm3: int = PALLET_CAP_CM3) -> dict:
    """{sku(str): unidades_por_palete(int)} — vazio se ficheiro indisponível.
    Por defeito lê da cache; com force=True relê o ficheiro (395k linhas, ~1-2 min)."""
    if CACHE.exists() and not force:
        try:
            return json.loads(CACHE.read_text(encoding="utf-8"))
        except Exception:
            pass
    if not force and not CACHE.exists():
        # primeira vez: tenta ler (é o que assegura a atualização automática)
        pass
    f = _find_file()
    if not f:
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
        ws = wb["VOLUMETRIA"] if "VOLUMETRIA" in wb.sheetnames else wb[wb.sheetnames[0]]
        lk = {}
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            try:
                sku = row[C_SKU]
                if sku is None:
                    continue
                sku = str(sku).strip()
                if sku.endswith(".0"):
                    sku = sku[:-2]
                if not sku or sku in lk:   # primeiro valor válido por SKU
                    continue
                vol = row[C_VOL]
                if not vol or float(vol) <= 0:
                    # fallback: volume = altura × comprimento × largura (cm)
                    _h, _l, _w = row[C_H], row[C_L], row[C_W]
                    vol = (float(_h) * float(_l) * float(_w)) if (_h and _l and _w) else None
                upp = units_per_pallet_from_volume(vol, cap_cm3)
                if upp:
                    lk[sku] = upp
            except (IndexError, TypeError, ValueError):
                continue
        wb.close()
        CACHE.parent.mkdir(parents=True, exist_ok=True)
        CACHE.write_text(json.dumps(lk, ensure_ascii=False), encoding="utf-8")
        return lk
    except Exception:
        return {}
