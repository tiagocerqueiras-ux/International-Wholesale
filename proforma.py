"""
proforma.py — Gerador de Proforma Invoice (.xlsx)
==================================================
Gera uma Proforma Invoice profissional a partir dos dados do deal + cliente.
Preenche automaticamente:
  - Dados do vendedor (Silly Bee, LDA — estático)
  - Dados do comprador (BD de clientes)
  - Referência e data (do deal)
  - Linhas de produto (SKU, EAN, Model, Artigo, Preço Unit., Qty, Total)
  - Totais (Subtotal, IVA, Total, Qty Total)
  - Condições comerciais (Pagamento, Incoterm, Método)
  - Dados bancários (IBAN/SWIFT — estático Silly Bee)

Usage:
    from proforma import generate_proforma
    xlsx_bytes = generate_proforma(deal, client)
    # deal: dict do get_deal()  |  client: dict do get_client()
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Identidade do Vendedor — escolha Worten / Silly Bee ──────────────────────
WORTEN = {
    "name":       "WORTEN EQUIPAMENTOS PARA O LAR, SA",
    "addr1":      "Rua João Mendonça, 529",
    "addr2":      "Senhora da Hora",
    "addr3":      "4464-501 Senhora da Hora · Portugal",
    "cap_social": "Cap. Social: 1 400 000 EUR",
    "vat":        "NIF/VAT: 503630330",
    "reg_com":    "Cons. Reg. Com. Porto sob o nº 503630330",
    "sirpeee":    "Nº SIRPEEE PT001257",
    "email":      "",
    "phone":      "",
    "bank":       "Banco Santander Totta",
    "nib":        "0018 6484 0200 0062 8998 2",
    "iban":       "PT50 0018 6484 0200 0062 8998 2",
    "swift":      "TOTAPTPL",
}

SILLY_BEE = {
    "name":       "SILLY BEE, LDA",
    "addr1":      "Rua Manuel Ferreira de Andrade, 10 – 9B",
    "addr2":      "São Domingos de Benfica",
    "addr3":      "1500-417 Lisboa · Portugal",
    "cap_social": "",
    "vat":        "NIF/VAT: 518 674 134",
    "reg_com":    "",
    "sirpeee":    "",
    "email":      "tiago.cerqueira@transglobalchain.com",
    "phone":      "+351 919 540 175",
    "bank":       "Banco CGD",
    "nib":        "",
    "iban":       "PT50 0018 0003 6529 6626 0205 1",
    "swift":      "TOTAPTPL",
}

SELLERS = {"Worten": WORTEN, "Silly Bee": SILLY_BEE}
SELLER = WORTEN   # default (retrocompatibilidade)

# ── Paleta ───────────────────────────────────────────────────────────────────
NAVY   = "1B2744"
GOLD   = "C49A3C"
LIGHT  = "F5F6FA"
WHITE  = "FFFFFF"
BORDER = "D0D4DE"
DARK_TEXT  = "1B2744"
LIGHT_TEXT = "FFFFFF"

# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color=DARK_TEXT, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_thin(sides="lrtb") -> Border:
    thin = Side(style="thin", color=BORDER)
    none = Side(style=None)
    return Border(
        left=thin   if "l" in sides else none,
        right=thin  if "r" in sides else none,
        top=thin    if "t" in sides else none,
        bottom=thin if "b" in sides else none,
    )

def _border_medium(sides="lrtb") -> Border:
    med  = Side(style="medium", color=NAVY)
    none = Side(style=None)
    return Border(
        left=med   if "l" in sides else none,
        right=med  if "r" in sides else none,
        top=med    if "t" in sides else none,
        bottom=med if "b" in sides else none,
    )

def _apply(cell, *, fill=None, font=None, align=None, border=None, fmt=None):
    if fill:   cell.fill      = fill
    if font:   cell.font      = font
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fmt:    cell.number_format = fmt


# ── Utilidades de dados ───────────────────────────────────────────────────────

def _num(v):
    try:
        if v is None: return 0.0
        if isinstance(v, str):
            v = v.replace("€","").replace(",",".").replace("%","").strip()
        return float(v)
    except (ValueError, TypeError):
        return 0.0

def _vat_rate_from_deal(deal: dict) -> float:
    vat_str = str(deal.get("IVA", "") or "")
    if "23" in vat_str: return 0.23
    if "13" in vat_str: return 0.13
    if "6"  in vat_str: return 0.06
    return 0.0

def _proforma_ref(deal: dict) -> str:
    did = str(deal.get("Deal ID", deal.get("deal_id", "")) or "")
    now = datetime.now()
    month_pt = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    m = month_pt[now.month - 1]
    y = str(now.year)[-2:]
    seq = did.split("-")[-1] if did else "001"
    return f"PROFORMA INVOICE: {m}{y}_I_Ref: {did or 'N/A'}"

def _client_lines(client: dict, deal: dict) -> dict:
    """Agrega os campos do comprador de várias fontes."""
    name    = (client.get("company_name") or deal.get("Cliente") or deal.get("client") or "").strip()
    contact = (client.get("contact_name") or "").strip()
    addr    = ""
    zip_    = ""
    city    = ""
    country = (client.get("country") or deal.get("País") or deal.get("country") or "").strip()
    vat     = (client.get("vat") or "").strip()
    email   = (client.get("billing_email") or client.get("contact_email")
               or deal.get("Email Cliente") or deal.get("client_email") or "").strip()

    # Morada de entrega — preferir a do cliente, depois do deal
    delivery = None
    da = client.get("delivery_addresses")
    if isinstance(da, list) and da:
        delivery = da[0]
    if delivery:
        addr    = delivery.get("address","")
        zip_    = delivery.get("zip","")
        city    = delivery.get("city","")
        country = delivery.get("country","") or country

    return dict(name=name, contact=contact, addr=addr, zip_=zip_,
                city=city, country=country, vat=vat, email=email)

def _build_lines(deal: dict) -> list[dict]:
    """Extrai linhas de produto do deal → lista de dicts."""
    skus = deal.get("_skus_detail") or deal.get("skus_detail") or {}
    lines = []
    for sku, info in skus.items():
        d    = info.get("data") or {}
        qty  = _num(info.get("qty", 1))
        pvp  = _num(info.get("pvp") or info.get("fc_final") or d.get("ufc_raw") or 0)
        name = f"{d.get('brand','')} {d.get('name','')}".strip()
        lines.append({
            "sku":   sku,
            "ean":   str(d.get("ean","") or ""),
            "model": str(d.get("model","") or ""),
            "name":  name or str(info.get("name","") or ""),
            "price": pvp,
            "qty":   qty,
            "total": round(pvp * qty, 3),
        })
    return lines


# ── Gerador principal ─────────────────────────────────────────────────────────

def generate_proforma(deal: dict, client: dict, seller: str = "Worten") -> bytes:
    """
    Devolve bytes de um .xlsx com a Proforma Invoice pronta a descarregar.
    deal   — resultado de get_deal()
    client — resultado de get_client() ou {} quando não disponível
    seller — "Worten" ou "Silly Bee" (empresa fornecedora)
    """
    SELLER = SELLERS.get(seller, WORTEN)
    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice"

    # ── Larguras de coluna  (A-I) ────────────────────────────────────────────
    col_widths = [2, 14, 16, 22, 30, 14, 10, 12, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Alturas de linha fixas ────────────────────────────────────────────────
    ws.row_dimensions[1].height  = 6    # topo margem
    ws.row_dimensions[2].height  = 50   # banner principal
    ws.row_dimensions[3].height  = 8    # espaço
    ws.row_dimensions[4].height  = 14
    ws.row_dimensions[5].height  = 12
    ws.row_dimensions[6].height  = 12
    ws.row_dimensions[7].height  = 12
    ws.row_dimensions[8].height  = 12
    ws.row_dimensions[9].height  = 12
    ws.row_dimensions[10].height = 8    # espaço
    ws.row_dimensions[11].height = 14   # título proforma ref
    ws.row_dimensions[12].height = 8    # espaço

    # ── ROW 1: espaço superior ────────────────────────────────────────────────
    for c in range(1, 10):
        ws.cell(row=1, column=c).fill = _fill(WHITE)

    # ── ROW 2: Banner (A2:I2) ────────────────────────────────────────────────
    ws.merge_cells("B2:E2")
    ws.merge_cells("F2:H2")

    banner_l = ws["B2"]
    banner_l.value     = SELLER["name"]
    banner_l.font      = Font(name="Calibri", bold=True, size=18, color=LIGHT_TEXT)
    banner_l.fill      = _fill(NAVY)
    banner_l.alignment = _align("left", "center")

    banner_r = ws["F2"]
    banner_r.value     = "PROFORMA INVOICE"
    banner_r.font      = Font(name="Calibri", bold=True, size=14, color=GOLD)
    banner_r.fill      = _fill(NAVY)
    banner_r.alignment = _align("right", "center")

    for c in [1, 9]:
        ws.cell(row=2, column=c).fill = _fill(NAVY)

    # gold accent strip (A3:I3)
    ws.merge_cells("A3:I3")
    ws["A3"].fill = _fill(GOLD)

    # ── ROW 4–9: Seller info (esquerda) + Buyer info (direita) ───────────────
    def _info_cell(row, col, val, bold=False, size=9):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=bold, size=size)
        c.alignment = _align("left", "center")
        c.fill      = _fill(WHITE)

    # Seller — 6 linhas (rows 4-9), campos do exemplo quando existem
    _seller_lines = [
        (SELLER["name"], True),
        (SELLER.get("addr1", ""), False),
        (" · ".join(filter(None, [SELLER.get("addr2", ""), SELLER.get("addr3", "")])), False),
        (" · ".join(filter(None, [SELLER.get("cap_social", ""), SELLER.get("vat", "")])), False),
        (SELLER.get("reg_com", ""), False),
        (SELLER.get("sirpeee") or SELLER.get("email") or "", False),
    ]
    for _i, (_val, _b) in enumerate(_seller_lines):
        _info_cell(4 + _i, 2, _val, bold=_b, size=10 if _b else 9)

    # Buyer — construir
    buyer = _client_lines(client, deal)

    # Label "BILL TO" com fundo navy
    ws.merge_cells("F4:H4")
    label = ws["F4"]
    label.value     = "BILL TO"
    label.font      = Font(name="Calibri", bold=True, size=9, color=LIGHT_TEXT)
    label.fill      = _fill(NAVY)
    label.alignment = _align("left", "center")

    def _buyer_cell(row, val):
        ws.merge_cells(f"F{row}:H{row}")
        c = ws.cell(row=row, column=6, value=val)
        c.font      = _font(size=9)
        c.alignment = _align("left", "center")
        c.fill      = _fill(LIGHT)

    _buyer_cell(5, buyer["name"])
    addr_line = " · ".join(filter(None, [buyer["addr"], buyer["zip_"], buyer["city"]]))
    _buyer_cell(6, addr_line or "—")
    _buyer_cell(7, buyer["country"])
    _buyer_cell(8, f"VAT: {buyer['vat']}" if buyer["vat"] else "")
    _buyer_cell(9, buyer["email"])

    # Fill lateral esquerda/direita rows 4-9
    for row in range(4, 10):
        ws.cell(row=row, column=1).fill = _fill(WHITE)
        ws.cell(row=row, column=9).fill = _fill(WHITE)
        for c in range(3, 6):
            ws.cell(row=row, column=c).fill = _fill(WHITE)

    # ── ROW 10: espaço ────────────────────────────────────────────────────────
    ws.merge_cells("A10:I10")
    ws["A10"].fill = _fill(WHITE)

    # ── ROW 11: Referência + Data ─────────────────────────────────────────────
    ref_str  = _proforma_ref(deal)
    date_val = datetime.now()

    ws.merge_cells("B11:E11")
    ref_cell = ws["B11"]
    ref_cell.value     = ref_str
    ref_cell.font      = Font(name="Calibri", bold=True, size=10, color=NAVY)
    ref_cell.alignment = _align("left", "center")
    ref_cell.fill      = _fill(WHITE)

    ws.merge_cells("F11:H11")
    date_cell = ws["F11"]
    date_cell.value          = date_val
    date_cell.number_format  = "DD/MM/YYYY"
    date_cell.font           = Font(name="Calibri", bold=True, size=10, color=NAVY)
    date_cell.alignment      = _align("right", "center")
    date_cell.fill           = _fill(WHITE)

    for c in [1, 9]:
        ws.cell(row=11, column=c).fill = _fill(WHITE)
    for c in range(3, 6):
        ws.cell(row=11, column=c).fill = _fill(WHITE)

    # ── ROW 12: espaço ────────────────────────────────────────────────────────
    ws.merge_cells("A12:I12")
    ws["A12"].fill = _fill(WHITE)

    # ── ROW 13: Cabeçalho tabela ──────────────────────────────────────────────
    HDR_ROW = 13
    ws.row_dimensions[HDR_ROW].height = 22
    headers = ["SKU", "EAN", "MODEL", "DESCRIPTION", "UNIT PRICE", "QTY", "TOTAL"]
    col_map = [2, 3, 4, 5, 6, 7, 8]  # colunas B-H

    for h_text, col in zip(headers, col_map):
        cell = ws.cell(row=HDR_ROW, column=col, value=h_text)
        cell.fill      = _fill(NAVY)
        cell.font      = Font(name="Calibri", bold=True, size=9, color=LIGHT_TEXT)
        cell.alignment = _align("center" if col != 5 else "left", "center")
        cell.border    = _border_medium("lrtb")
    ws.cell(row=HDR_ROW, column=1).fill = _fill(NAVY)
    ws.cell(row=HDR_ROW, column=9).fill = _fill(NAVY)

    # ── Linhas de produto ─────────────────────────────────────────────────────
    lines = _build_lines(deal)
    FIRST_DATA_ROW = HDR_ROW + 1

    for i, ln in enumerate(lines):
        row = FIRST_DATA_ROW + i
        ws.row_dimensions[row].height = 16
        bg = WHITE if i % 2 == 0 else LIGHT

        data_row = [ln["sku"], ln["ean"], ln["model"], ln["name"],
                    ln["price"], int(ln["qty"]), ln["total"]]

        for val, col in zip(data_row, col_map):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = _fill(bg)
            cell.border = _border_thin("lrtb")
            cell.font   = _font(size=9)
            right_cols = {6, 7, 8}
            cell.alignment = _align("right" if col in right_cols else "left", "center")
            if col in (6, 8):  # UNIT PRICE, TOTAL
                cell.number_format = '#,##0.000 "€"'
            if col == 7:       # QTY
                cell.number_format = '#,##0'

        ws.cell(row=row, column=1).fill = _fill(bg)
        ws.cell(row=row, column=9).fill = _fill(bg)

    # Linha de separação após produtos
    sep_row = FIRST_DATA_ROW + len(lines)
    ws.row_dimensions[sep_row].height = 3
    ws.merge_cells(f"A{sep_row}:I{sep_row}")
    ws.cell(row=sep_row, column=1).fill = _fill(GOLD)

    # ── Totais ────────────────────────────────────────────────────────────────
    subtotal   = sum(ln["total"] for ln in lines)
    vat_rate   = _vat_rate_from_deal(deal)
    vat_amount = round(subtotal * vat_rate, 3)
    total_amt  = round(subtotal + vat_amount, 3)
    total_qty  = int(sum(ln["qty"] for ln in lines))
    freight    = _num(deal.get("Frete (€)") or deal.get("freight_cost") or 0)

    TOT_START = sep_row + 1
    tot_rows = [
        ("Amount",       subtotal,   '#,##0.000 "€"'),
        ("VAT",          vat_amount, '#,##0.000 "€"'),
        ("Transport",    freight,    '#,##0.000 "€"') if freight else None,
        ("Total Amount", total_amt,  '#,##0.000 "€"'),
        ("Total Qty",    total_qty,  '#,##0'),
    ]
    tot_rows = [r for r in tot_rows if r is not None]

    for j, (label, val, fmt) in enumerate(tot_rows):
        row = TOT_START + j
        ws.row_dimensions[row].height = 16
        is_bold = label in {"Total Amount", "Total Qty"}
        is_grand = label == "Total Amount"

        # B-E: espaço vazio branco
        ws.merge_cells(f"B{row}:E{row}")
        ws.cell(row=row, column=2).fill = _fill(WHITE)
        for c in [1, 9]:
            ws.cell(row=row, column=c).fill = _fill(WHITE)

        # F-G: label (merge)
        ws.merge_cells(f"F{row}:G{row}")
        lbl_cell = ws.cell(row=row, column=6, value=label)
        lbl_cell.font      = _font(bold=is_bold, size=9)
        lbl_cell.alignment = _align("right", "center")
        lbl_cell.fill      = _fill(LIGHT)
        lbl_cell.border    = _border_thin("lrtb")

        # H: valor
        val_cell = ws.cell(row=row, column=8, value=val)
        val_cell.font          = _font(bold=is_bold, size=9,
                                       color=NAVY if is_grand else DARK_TEXT)
        val_cell.alignment     = _align("right", "center")
        val_cell.fill          = _fill("E8F0FC" if is_grand else LIGHT)
        val_cell.border        = _border_thin("lrtb")
        val_cell.number_format = fmt

    # ── Condições Comerciais ──────────────────────────────────────────────────
    COND_ROW = TOT_START + len(tot_rows) + 2
    ws.row_dimensions[COND_ROW - 1].height = 8
    ws.row_dimensions[COND_ROW].height     = 22

    ws.merge_cells(f"B{COND_ROW}:H{COND_ROW}")
    hdr = ws.cell(row=COND_ROW, column=2, value="COMMERCIAL TERMS")
    hdr.font      = Font(name="Calibri", bold=True, size=9, color=LIGHT_TEXT)
    hdr.fill      = _fill(NAVY)
    hdr.alignment = _align("left", "center")
    hdr.border    = _border_thin("lrtb")
    for c in [1, 9]:
        ws.cell(row=COND_ROW, column=c).fill = _fill(NAVY)

    payment  = str(deal.get("Pagamento", deal.get("payment_conditions","")) or "In Advance")
    incoterm = str(deal.get("Incoterm",  deal.get("incoterm",""))            or "DAP")
    method   = "Bank Transfer"

    cond_items = [
        ("Payment Terms",  payment),
        ("Incoterm",       incoterm),
        ("Payment Method", method),
    ]

    for k, (label, value) in enumerate(cond_items):
        r = COND_ROW + 1 + k
        ws.row_dimensions[r].height = 14

        lbl = ws.cell(row=r, column=2, value=label + ":")
        lbl.font      = _font(bold=True, size=9)
        lbl.alignment = _align("left", "center")
        lbl.fill      = _fill(LIGHT)
        lbl.border    = _border_thin("lrtb")

        ws.merge_cells(f"C{r}:H{r}")
        val_c = ws.cell(row=r, column=3, value=value)
        val_c.font      = _font(size=9)
        val_c.alignment = _align("left", "center")
        val_c.fill      = _fill(WHITE)
        val_c.border    = _border_thin("lrtb")

        for c in [1, 9]:
            ws.cell(row=r, column=c).fill = _fill(WHITE)

    # ── Dados Bancários ───────────────────────────────────────────────────────
    BANK_ROW = COND_ROW + len(cond_items) + 2
    ws.row_dimensions[BANK_ROW - 1].height = 8
    ws.row_dimensions[BANK_ROW].height     = 22

    ws.merge_cells(f"B{BANK_ROW}:H{BANK_ROW}")
    bank_hdr = ws.cell(row=BANK_ROW, column=2, value="BANK DETAILS")
    bank_hdr.font      = Font(name="Calibri", bold=True, size=9, color=LIGHT_TEXT)
    bank_hdr.fill      = _fill(NAVY)
    bank_hdr.alignment = _align("left", "center")
    bank_hdr.border    = _border_thin("lrtb")
    for c in [1, 9]:
        ws.cell(row=BANK_ROW, column=c).fill = _fill(NAVY)

    bank_items = [
        ("Beneficiary", SELLER["name"]),
        ("Bank",        SELLER.get("bank", "")),
        ("NIB",         SELLER.get("nib", "")),
        ("IBAN",        SELLER.get("iban", "")),
        ("SWIFT / BIC", SELLER.get("swift", "")),
    ]
    bank_items = [(_l, _v) for _l, _v in bank_items if _v]

    for k, (label, value) in enumerate(bank_items):
        r = BANK_ROW + 1 + k
        ws.row_dimensions[r].height = 14

        lbl = ws.cell(row=r, column=2, value=label + ":")
        lbl.font      = _font(bold=True, size=9)
        lbl.alignment = _align("left", "center")
        lbl.fill      = _fill(LIGHT)
        lbl.border    = _border_thin("lrtb")

        ws.merge_cells(f"C{r}:H{r}")
        val_c = ws.cell(row=r, column=3, value=value)
        val_c.font      = _font(size=9)
        val_c.alignment = _align("left", "center")
        val_c.fill      = _fill(WHITE)
        val_c.border    = _border_thin("lrtb")

        for c in [1, 9]:
            ws.cell(row=r, column=c).fill = _fill(WHITE)

    # ── Rodapé ────────────────────────────────────────────────────────────────
    FOOT_ROW = BANK_ROW + len(bank_items) + 2
    ws.row_dimensions[FOOT_ROW - 1].height = 6
    ws.row_dimensions[FOOT_ROW].height     = 18

    ws.merge_cells(f"A{FOOT_ROW}:I{FOOT_ROW}")
    foot = ws["A" + str(FOOT_ROW)]
    foot.fill = _fill(GOLD)

    ws.row_dimensions[FOOT_ROW + 1].height = 14
    ws.merge_cells(f"B{FOOT_ROW+1}:H{FOOT_ROW+1}")
    note = ws.cell(row=FOOT_ROW + 1, column=2,
                   value="This proforma invoice is not a tax document. Valid for 30 days from the date of issue.")
    note.font      = Font(name="Calibri", italic=True, size=8, color="888888")
    note.alignment = _align("center", "center")
    note.fill      = _fill(WHITE)

    # ── Remover gridlines ─────────────────────────────────────────────────────
    ws.sheet_view.showGridLines = False

    # ── Print setup ───────────────────────────────────────────────────────────
    ws.page_setup.orientation    = "portrait"
    ws.page_setup.paperSize      = 9   # A4
    ws.page_setup.fitToWidth     = 1
    ws.page_setup.fitToHeight    = 0
    ws.page_margins.left         = 0.5
    ws.page_margins.right        = 0.5
    ws.page_margins.top          = 0.5
    ws.page_margins.bottom       = 0.5

    # ── Serializar para bytes ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
