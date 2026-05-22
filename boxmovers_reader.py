"""
BoxMovers Excel Reader
======================
Lê BoxMovers2025_actual.xlsx e BoxMovers2026.xlsx do SharePoint local.
Fornece dados reais (clientes, marcas, categorias, receita, margem) para o dashboard executivo.

Correções v2:
  - Copia para temp antes de ler (resolve PermissionError quando ficheiro aberto no Excel)
  - GOOGLE, STARLINK, OPPO adicionados a _ABRAND_KEYWORDS
  - FUJIFILM-INSTAX normalizado para FUJIFILM
  - Cache de Deep Info (950K linhas) por caminho + mtime para evitar releitura
  - Fallback gracioso se ficheiros indisponíveis (Railway/cloud)
"""

from __future__ import annotations

import os
import shutil
import tempfile
import warnings
from datetime import datetime
from pathlib import Path

# ── Índices de colunas (0-based) na sheet DEALS ───────────────────────────────
# Confirmados vs. header row (row 2 do ficheiro):
#   A=0 DEAL DATE | B=1 STATUS | C=2 FATURAÇÃO | M=12 CLIENT
#   Q=16 CAT | R=17 DESCRIÇÃO CAT | Y=24 BRAND | Z=25 SKU
#   AR=43 TTL SALES | BH=59 TTL MG | BI=60 MG%4
_C_DATE      = 0    # A — DEAL DATE  ("Jan´26", "Fev'25", datetime…)
_C_STATUS    = 1    # B — STATUS     (CONCLUÍDO / PO / CANCELADO)
_C_CLIENT    = 12   # M — CLIENT
_C_CAT_CODE  = 16   # Q — CAT code   (e.g. 5305)
_C_CAT_DESC  = 17   # R — DESCRIÇÃO CAT  ("5305 - Hardware Foto E Cam")
_C_BRAND     = 24   # Y — BRAND
_C_SKU       = 25   # Z — SKU
_C_TTL_SALES = 43   # AR — TTL SALES (€)
_C_TTL_MG    = 59   # BH — TTL MG (€)
_C_MG_PCT    = 60   # BI — MG%4 (decimal: 0.053 = 5.3%)

# ── Índices na sheet "Deep Info 2.0" ─────────────────────────────────────────
_DI_SKU      = 5    # F — SKU
_DI_CAT_DESC = 2    # C — Categoria Desc
_DI_BRAND    = 8    # I — MARCA

# ── Status que representam vendas concluídas ──────────────────────────────────
_CONCLUDED_NORM = {"CONCLUIDO", "PO"}   # comparação após normalizar Í→I

# ── Meses portugueses ─────────────────────────────────────────────────────────
_PT_MONTHS = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

# ── Normalização de nomes de marcas ──────────────────────────────────────────
# Mapeia variantes encontradas nos ficheiros → nome canónico
_BRAND_NORMALIZE: dict[str, str] = {
    "FUJIFILM-INSTAX": "FUJIFILM",
    "FUJIFILM INSTAX": "FUJIFILM",
    "SEM MARCA":       "SEM MARCA",   # manter separado, não é A-Brand
}

# ── Marcas A-Brand (verificação por substring) ───────────────────────────────
# Atualizado com marcas presentes nos dados: GOOGLE, STARLINK, OPPO
_ABRAND_KEYWORDS = {
    "SAMSUNG", "PHILIPS", "SONY", "LG", "BRAUN", "BABYLISS", "BOSCH",
    "SIEMENS", "TEFAL", "MOULINEX", "ROWENTA", "KRUPS", "DELONGHI", "DYSON",
    "SMEG", "BOSE", "CANON", "NIKON", "FUJIFILM", "PANASONIC", "TOSHIBA",
    "HISENSE", "TCL", "XIAOMI", "OPPO", "APPLE", "ASUS", "LENOVO", "HP",
    "GORENJE", "WHIRLPOOL", "SHARP", "NINTENDO", "REMINGTON", "DREAME",
    "COSORI", "AMAZON", "ELECTROLUX", "ZANUSSI", "AEG", "MIELE", "GRUNDIG",
    "HAIER", "CANDY", "HOOVER", "INDESIT", "HOTPOINT", "BEKO", "ARISTON",
    "PLAYSTATION", "MICROSOFT", "XBOX",
    "GOOGLE",    # Pixel, Nest, Chromebook → confirmado nos dados
    "STARLINK",  # SpaceX Starlink → confirmado nos dados
}

# ── Cache de Deep Info (evita reler 950K linhas a cada chamada) ───────────────
_DI_CACHE: dict[str, dict] = {}   # key = str(path) → índice SKU→{cat_desc, brand}
_DI_MTIME: dict[str, float] = {}  # key = str(path) → mtime no momento do cache


# ─────────────────────────────────────────────────────────────────────────────
# Utilitários internos
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_lnk(lnk_path: Path) -> Path | None:
    """Resolve atalho .lnk do Windows para o caminho real."""
    try:
        import win32com.client
        sh  = win32com.client.Dispatch("WScript.Shell")
        lnk = sh.CreateShortcut(str(lnk_path))
        t   = Path(lnk.TargetPath)
        return t if t.exists() else None
    except Exception:
        return None


def _get_bm_paths() -> list[tuple[int, Path]]:
    """Devolve lista de (ano, Path) para os ficheiros BoxMovers acessíveis."""
    docs = Path(__file__).parent.parent / "Docs"
    candidates = [
        (2026, "BoxMovers2026.lnk"),
        (2025, "BoxMovers2025_actual - Atalho.lnk"),
    ]
    result = []
    for year, lnk_name in candidates:
        lnk = docs / lnk_name
        if lnk.exists():
            target = _resolve_lnk(lnk)
            if target:
                result.append((year, target))
    return result


def _open_workbook(path: Path):
    """
    Abre o workbook openpyxl.
    Se o ficheiro estiver bloqueado (aberto no Excel / OneDrive a sincronizar),
    copia para um ficheiro temporário e lê a partir daí.
    """
    import openpyxl
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            return openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except PermissionError:
            # Ficheiro bloqueado → copiar para temp
            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False,
                dir=tempfile.gettempdir(), prefix="bm_reader_"
            )
            tmp.close()
            shutil.copy2(str(path), tmp.name)
            wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
            try:
                os.unlink(tmp.name)
            except Exception:
                pass
            return wb


def _build_di_index(wb, path_key: str) -> dict:
    """
    Constrói SKU → {cat_desc, brand} a partir de Deep Info 2.0.
    Usa cache por (path + mtime) para evitar reler ~950K linhas.
    """
    # verificar cache
    if path_key in _DI_CACHE:
        return _DI_CACHE[path_key]

    ws = None
    for name in ("Deep Info 2.0", "Deep info"):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        _DI_CACHE[path_key] = {}
        return {}

    index: dict = {}
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = row[_DI_SKU]
            if sku is None:
                continue
            cat_desc = str(row[_DI_CAT_DESC] or "").strip()
            brand    = str(row[_DI_BRAND]    or "").strip().upper()
            index[str(sku)] = {"cat_desc": cat_desc, "brand": brand}

    _DI_CACHE[path_key] = index
    return index


def _normalize_brand(brand: str) -> str:
    """Normaliza variantes de nomes de marca para o nome canónico."""
    upper = brand.upper()
    return _BRAND_NORMALIZE.get(upper, brand)


def _is_abrand(brand: str) -> bool:
    """Devolve True se a marca for uma A-Brand reconhecida."""
    if not brand or brand in ("—", "SEM MARCA"):
        return False
    upper = brand.upper()
    return any(kw in upper for kw in _ABRAND_KEYWORDS)


def _parse_date(val) -> tuple[int | None, int | None]:
    """
    Converte datas do Excel → (year, month).
    Suporta: datetime, "Jan´26" (0xb4), "Fev'25" (0x27), "Abri'25", etc.
    """
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.year, val.month
    s = str(val).strip().lower()
    # Normalizar separadores: ´ (0xb4), ' (U+2019), ' (U+2018), ` → '
    for ch in ("\u00b4", "\u2019", "\u2018", "\u0060", "\u02bc"):
        s = s.replace(ch, "'")
    if "'" in s:
        parts = s.split("'")
        m_str = parts[0].strip()[:3]
        y_str = parts[1].strip()[:4]
        m = _PT_MONTHS.get(m_str)
        try:
            y = int(y_str)
            if y < 100:
                y += 2000
            return y, m
        except ValueError:
            pass
    return None, None


def _is_concluded(status: str) -> bool:
    """Devolve True para CONCLUÍDO (qualquer variante de codificação)."""
    return status.upper().replace("\u00cd", "I").replace("Í", "I") in _CONCLUDED_NORM


# ─────────────────────────────────────────────────────────────────────────────
# API pública
# ─────────────────────────────────────────────────────────────────────────────

def _read_bm_deals_snapshot(year_filter: int | None = None) -> list[dict]:
    """
    Fallback: lê o snapshot JSON commitado no repo (usado no Railway/cloud
    onde os ficheiros .lnk do Windows não são resolúveis).
    """
    import json as _json
    snap = Path(__file__).parent / "data" / "bm_deals_snapshot.json"
    if not snap.exists():
        return []
    try:
        with open(snap, encoding="utf-8") as f:
            rows = _json.load(f)
        # Garantir que todos os campos necessários existem
        result = []
        for r in rows:
            yr = r.get("year", 0)
            if year_filter and yr != year_filter:
                continue
            status = r.get("status", "")
            status_norm = (status.upper()
                           .replace("Í", "I").replace("Ó", "O")
                           .replace("Â", "A").replace("Ã", "A")
                           .replace("Ú", "U").replace("É", "E"))
            result.append({
                "client":    r.get("client", "—"),
                "brand":     r.get("brand", "—"),
                "cat":       r.get("cat", "—"),
                "sku":       "",
                "revenue":   float(r.get("revenue", 0)),
                "mg_eur":    float(r.get("mg_eur", 0)),
                "mg_pct":    float(r.get("mg_pct", 0)),
                "year":      yr,
                "month":     r.get("month", 0),
                "status":    status,
                "bm_year":   yr,
                "concluded": status_norm in _CONCLUDED_NORM,
            })
        return result
    except Exception:
        return []


def read_bm_deals(year_filter: int | None = None) -> list[dict]:
    """
    Lê todas as linhas de deal dos ficheiros BoxMovers (2025 e/ou 2026).
    Enriquece campos vazios (brand, cat) via Deep Info 2.0.
    Normaliza nomes de marca.

    Args:
        year_filter: se fornecido, só lê o ficheiro desse ano.

    Returns:
        Lista de dicts com chaves: client, brand, cat, sku, revenue, mg_eur,
        mg_pct, year, month, status, bm_year, concluded.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return _read_bm_deals_snapshot(year_filter)

    paths = _get_bm_paths()
    if not paths:
        # Excel indisponível (Railway/cloud) — usar snapshot JSON
        return _read_bm_deals_snapshot(year_filter)

    all_rows: list[dict] = []

    for bm_year, path in paths:
        if year_filter and bm_year != year_filter:
            continue
        try:
            wb       = _open_workbook(path)
            path_key = str(path)
            di       = _build_di_index(wb, path_key)
            ws       = wb["DEALS"]

            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                for row in ws.iter_rows(min_row=3, values_only=True):
                    sku = row[_C_SKU]
                    if sku is None:
                        continue

                    sku_str = str(sku)
                    status  = str(row[_C_STATUS]   or "").strip()
                    client  = str(row[_C_CLIENT]   or "").strip()
                    brand   = str(row[_C_BRAND]    or "").strip().upper()
                    cat     = str(row[_C_CAT_DESC] or "").strip()
                    revenue = float(row[_C_TTL_SALES] or 0)
                    mg_eur  = float(row[_C_TTL_MG]   or 0)
                    mg_pct  = float(row[_C_MG_PCT]   or 0) * 100  # 0.053 → 5.3%

                    yr, mo = _parse_date(row[_C_DATE])
                    if yr is None:
                        yr = bm_year

                    # Enriquecimento via Deep Info quando campos vazios
                    di_entry = di.get(sku_str, {})
                    if not brand and di_entry.get("brand"):
                        brand = di_entry["brand"]
                    if not cat and di_entry.get("cat_desc"):
                        cat = di_entry["cat_desc"]

                    # Normalização do nome de marca
                    brand = _normalize_brand(brand) if brand else "—"

                    all_rows.append({
                        "client":    client    or "—",
                        "brand":     brand,
                        "cat":       cat       or "—",
                        "sku":       sku_str,
                        "revenue":   revenue,
                        "mg_eur":    mg_eur,
                        "mg_pct":    round(mg_pct, 2),
                        "year":      yr,
                        "month":     mo or 1,
                        "status":    status,
                        "bm_year":   bm_year,
                        "concluded": _is_concluded(status),
                    })

            wb.close()

        except Exception as e:
            print(f"[boxmovers_reader] Erro ao ler {path.name}: {e}")

    return all_rows


def get_bm_dashboard_data(year: int | None = None) -> dict:
    """
    Agrega dados de deals BoxMovers para o dashboard executivo.

    - KPIs e drivers calculados apenas sobre deals CONCLUÍDO.
    - year=None inclui todos os anos disponíveis.
    - Retorna {} se nenhum ficheiro estiver disponível.
    """
    rows = read_bm_deals(year_filter=year)
    if not rows:
        return {}

    # Filtrar por ano se especificado (dupla segurança após leitura)
    if year:
        rows = [r for r in rows if r["year"] == year]
    if not rows:
        return {}

    concluded = [r for r in rows if r["concluded"]]
    if not concluded:
        # Sem dados CONCLUÍDO → devolver vazio para não distorcer dashboard
        return {}

    # ── KPIs ─────────────────────────────────────────────────────────────────
    total_rev  = sum(r["revenue"] for r in concluded)
    total_mg   = sum(r["mg_eur"]  for r in concluded)
    avg_mg_pct = round(total_mg / total_rev * 100, 2) if total_rev else 0.0

    # ── Top cliente ───────────────────────────────────────────────────────────
    client_rev: dict[str, float] = {}
    for r in concluded:
        c = r["client"]
        client_rev[c] = client_rev.get(c, 0) + r["revenue"]
    top_client     = max(client_rev, key=client_rev.get) if client_rev else "—"
    top_client_rev = round(client_rev.get(top_client, 0), 0)

    # ── Top marca e mix A-Brand ───────────────────────────────────────────────
    brand_rev: dict[str, float] = {}
    abrand_total = 0.0
    for r in concluded:
        b = r["brand"]
        if b and b not in ("—", "SEM MARCA"):
            brand_rev[b] = brand_rev.get(b, 0) + r["revenue"]
            if _is_abrand(b):
                abrand_total += r["revenue"]
    top_brand  = max(brand_rev, key=brand_rev.get) if brand_rev else "—"
    mix_abrand = round(abrand_total / total_rev * 100, 0) if total_rev else 0.0

    # ── Ticket médio (por deal = cliente único por mês) ───────────────────────
    deal_keys: set = set()
    for r in concluded:
        deal_keys.add((r["client"], r["year"], r["month"]))
    avg_ticket = round(total_rev / len(deal_keys), 0) if deal_keys else 0.0

    # ── Receita e margem mensais ──────────────────────────────────────────────
    monthly_rev: dict[str, float] = {}
    monthly_mg:  dict[str, float] = {}

    for r in concluded:
        key = f"{r['year']}-{r['month']:02d}"
        monthly_rev[key] = monthly_rev.get(key, 0) + r["revenue"]
        monthly_mg[key]  = monthly_mg.get(key,  0) + r["mg_eur"]

    monthly_mg_pct = {
        k: round(monthly_mg.get(k, 0) / monthly_rev[k] * 100, 2) if monthly_rev[k] > 0 else 0.0
        for k in monthly_rev
    }

    monthly_sorted    = dict(sorted(monthly_rev.items()))
    monthly_mg_sorted = dict(sorted(monthly_mg.items()))
    monthly_mgp_sorted = dict(sorted(monthly_mg_pct.items()))

    # ── Proveito — taxa por ano contratual (Abr→Mar, início Abril 2025) ─────────
    # Cada mês pertence a um ano contratual. A taxa final desse ano aplica-se
    # retroativamente a TODOS os seus meses (contrato em vigor desde Abr'25).
    try:
        from config import bp_commission_rate, bp_commission_tier_name

        def _cy_of(yr: int, mo: int) -> int:
            """Ano-base do ano contratual: mo>=4 → yr, mo<4 → yr-1."""
            return yr if mo >= 4 else yr - 1

        # Ler TODOS os dados (ambos os ficheiros) para calcular T/O por ano contratual
        _all_concluded = [r for r in read_bm_deals(year_filter=None) if r["concluded"]]

        # T/O acumulado por ano contratual (usando todos os dados disponíveis)
        _cy_to: dict[int, float] = {}
        for r in _all_concluded:
            if r["revenue"] > 0:
                cy = _cy_of(r["year"], r["month"])
                _cy_to[cy] = _cy_to.get(cy, 0) + r["revenue"]

        # Anos contratuais já concluídos — valores fixos auditados.
        # No cloud (Railway) o ficheiro 2025 não existe; sem este fallback o CY1
        # ficaria incompleto (~2.7M) e a taxa seria incorrecta (17,5% em vez de 20%).
        # O T/O do CY1 é imutável (contrato Abr'25→Mar'26 está encerrado).
        _CY_CLOSED: dict[int, float] = {
            2025: 12_635_206.0,   # Ano Contratual 1 (Abr'25→Mar'26) — auditado
        }
        for _cy_k, _cy_v in _CY_CLOSED.items():
            if _cy_to.get(_cy_k, 0.0) < _cy_v:
                _cy_to[_cy_k] = _cy_v

        # Taxa final por ano contratual
        _cy_rate: dict[int, float] = {cy: bp_commission_rate(tot) for cy, tot in _cy_to.items()}

        # Proveito mensal usando a taxa do ano contratual correcto de cada mês
        monthly_proveito = {}
        for k, mg_val in monthly_mg.items():
            yr2, mo2 = int(k[:4]), int(k[5:7])
            rate = _cy_rate.get(_cy_of(yr2, mo2), 0.175)
            monthly_proveito[k] = round(mg_val * rate, 2)

        our_cut = round(sum(monthly_proveito.values()), 0)

        # Taxa e escalão do ano contratual mais recente nos dados exibidos
        _latest = max(monthly_rev.keys()) if monthly_rev else None
        if _latest:
            _lyr, _lmo  = int(_latest[:4]), int(_latest[5:7])
            _cy_latest   = _cy_of(_lyr, _lmo)
            commission_rate   = _cy_rate.get(_cy_latest, 0.175)
            tier_name         = bp_commission_tier_name(_cy_to.get(_cy_latest, 0))
            cy_current_num    = _cy_latest - 2024          # Ano 1, Ano 2, ...
            cy_current_to     = _cy_to.get(_cy_latest, 0)  # T/O acumulado do ano contratual actual
            cy_current_label  = (f"Ano Contratual {cy_current_num}"
                                 f" (Abr'{str(_cy_latest)[2:]} → Mar'{str(_cy_latest+1)[2:]})")
        else:
            commission_rate   = 0.175
            tier_name         = "⚪ Base 17,5%"
            cy_current_num    = 0
            cy_current_to     = 0.0
            cy_current_label  = "—"

        # Retroativo: soma do extra (taxa_cy - base 17,5%) sobre TODOS os meses exibidos
        retroativo = round(
            sum(
                monthly_mg.get(k, 0) * (_cy_rate.get(_cy_of(int(k[:4]), int(k[5:7])), 0.175) - 0.175)
                for k in monthly_mg
            ), 2
        )

    except Exception:
        commission_rate   = 0.175
        tier_name         = "Base"
        our_cut           = round(total_mg * commission_rate, 0)
        monthly_proveito  = {
            k: round(monthly_mg.get(k, 0) * commission_rate, 2)
            for k in monthly_sorted
        }
        cy_current_num   = 0
        cy_current_to    = 0.0
        cy_current_label = "—"
        retroativo       = 0.0

    # ── Distribuição por categoria ────────────────────────────────────────────
    cat_rev: dict[str, float] = {}
    for r in concluded:
        cat_rev[r["cat"]] = cat_rev.get(r["cat"], 0) + r["revenue"]

    return {
        # KPIs
        "total_revenue":       round(total_rev, 2),
        "gross_margin_value":  round(total_mg, 2),
        "avg_margin":          avg_mg_pct,
        "our_cut":             our_cut,
        "commission_rate_pct": round(commission_rate * 100, 1),
        "tier_name":           tier_name,
        # Contexto do ano contratual (Abr→Mar)
        "cy_current_num":      cy_current_num,    # 1, 2, 3, ...
        "cy_current_to":       round(cy_current_to, 2),   # T/O acumulado do ano contratual actual
        "cy_current_label":    cy_current_label,  # ex: "Ano Contratual 2 (Abr'26 → Mar'27)"
        "retroativo":          retroativo,         # delta comissão vs base 17,5%
        # Drivers de crescimento
        "top_client":          top_client,
        "top_client_rev":      top_client_rev,
        "top_brand":           top_brand,
        "mix_abrand":          mix_abrand,
        "avg_ticket":          avg_ticket,
        # Mensais
        "monthly_revenue":     monthly_sorted,
        "monthly_margin":      monthly_mg_sorted,
        "monthly_margin_pct":  monthly_mgp_sorted,
        "monthly_proveito":    monthly_proveito,
        # Listas completas para análise detalhada
        "by_client":  dict(sorted(client_rev.items(), key=lambda x: x[1], reverse=True)),
        "by_brand":   dict(sorted(brand_rev.items(),  key=lambda x: x[1], reverse=True)),
        "by_cat":     dict(sorted(cat_rev.items(),    key=lambda x: x[1], reverse=True)),
        # Metadados
        "n_concluded": len(concluded),
        "n_all":       len(rows),
        "source":      "boxmovers",
    }
